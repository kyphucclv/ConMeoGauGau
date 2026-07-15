"""Whitelisted canonical reporting queries for the app UI."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from dataclasses import dataclass
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from db import fetch_all


@dataclass(frozen=True)
class Report:
    key: str
    label: str
    query: str
    metric_keys: tuple[str, ...] = ()


REPORTS: tuple[Report, ...] = (
    Report(
        "cohort_dashboard",
        "Cohort/course-run dashboard",
        """
        SELECT class_code, course_code, course_name, run_number, course_run_status,
               enrollment_count, active_enrollments, completed_enrollments,
               completed_meetings, cancelled_meetings, non_cancelled_units,
               average_attendance_ratio, exam_eligible_enrollments
        FROM v_cohort_course_run_dashboard
        ORDER BY class_code, course_name, run_number
        LIMIT 300
        """,
        ("attendance_ratio", "effective_exam_eligible"),
    ),
    Report(
        "current_employee_state",
        "Current employee state",
        """
        SELECT emp_code, full_name, employment_status, business_unit_name,
               job_role_name, class_code, course_code, course_name,
               enrollment_status, course_run_status
        FROM v_current_employee_state
        ORDER BY full_name
        LIMIT 300
        """,
    ),
    Report(
        "attendance_eligibility",
        "Attendance and eligibility",
        """
        SELECT employee_id, course_run_id, enrollment_status, start_session_number,
               applicable_units, present_units, absent_units, makeup_present_units,
               attendance_ratio, calculated_exam_eligible, effective_exam_eligible,
               exam_eligibility_override
        FROM v_run_enrollment_attendance
        ORDER BY course_run_id, employee_id
        LIMIT 300
        """,
        ("attendance_ratio", "effective_exam_eligible"),
    ),
    Report(
        "monthly_sessions",
        "Monthly session units",
        """
        SELECT session_month, course_run_id, cohort_id, course_id,
               credited_session_units, final_test_units, final_test_duration_minutes
        FROM v_monthly_session_units
        ORDER BY session_month DESC, course_run_id
        LIMIT 300
        """,
        ("sessions_per_month",),
    ),
    Report(
        "progress_summary",
        "Progress summary",
        """
        SELECT emp_code, full_name, entrance_level_name, current_level_name,
               highest_level_name, current_progress, peak_progress, regression_flag
        FROM v_employee_progress_summary
        ORDER BY full_name
        LIMIT 300
        """,
        ("current_level", "highest_level", "current_progress", "peak_progress", "regression_flag"),
    ),
    Report(
        "historical_enrollment_snapshot",
        "Historical enrollment snapshot",
        """
        SELECT emp_code, full_name, class_code, course_code, course_name,
               enrollment_status, enrollment_business_unit, enrollment_job_role,
               start_session_number, transfer_from_enrollment_id
        FROM v_historical_enrollment_snapshot
        ORDER BY enrollment_created_at DESC
        LIMIT 300
        """,
    ),
    Report(
        "unresolved_quality_issues",
        "Unresolved quality issues",
        """
        SELECT source, issue_code, entity_type, entity_key, source_sheet,
               source_row_number, created_at
        FROM v_unresolved_quality_issues
        ORDER BY created_at DESC
        LIMIT 300
        """,
        ("unresolved_quality_issues",),
    ),
)


def report_by_label(label: str) -> Report:
    for report in REPORTS:
        if report.label == label:
            return report
    raise KeyError(label)


def run_report(pool, report: Report):
    return fetch_all(pool, report.query)


def metric_definitions(pool, metric_keys: Iterable[str]):
    keys = list(metric_keys)
    if not keys:
        return []
    placeholders = ", ".join(["%s"] * len(keys))
    return fetch_all(
        pool,
        f"""
        SELECT metric_key, metric_name, definition, numerator_definition, denominator_definition
        FROM v_reporting_metric_definitions
        WHERE metric_key IN ({placeholders})
        ORDER BY metric_key
        """,
        keys,
    )


def monthly_review_data(pool, review_month: date) -> dict[str, object]:
    """Canonical, inspectable detail tables for one owner-defined review month."""
    next_month = date(review_month.year + (review_month.month == 12), (review_month.month % 12) + 1, 1)
    program = fetch_all(pool, """
        SELECT c.class_code, co.course_code, co.course_name, cr.course_run_id,
               count(DISTINCT re.run_enrollment_id) FILTER (WHERE re.status='active') AS active_participants,
               count(DISTINCT re.run_enrollment_id) FILTER (WHERE re.status='active' AND lifetime.enrollment_count >= 2) AS repeated_participants,
               count(DISTINCT su.sequence_in_run) FILTER (WHERE m.status='planned') AS planned_sessions,
               count(DISTINCT su.sequence_in_run) FILTER (WHERE m.status='completed') AS delivered_sessions
        FROM course_runs cr JOIN cohorts c ON c.cohort_id=cr.cohort_id JOIN courses co ON co.course_id=cr.course_id
        LEFT JOIN run_enrollments re ON re.course_run_id=cr.course_run_id
        LEFT JOIN (SELECT employee_id,count(*) AS enrollment_count FROM run_enrollments GROUP BY employee_id) lifetime ON lifetime.employee_id=re.employee_id
        LEFT JOIN session_units su ON su.course_run_id=cr.course_run_id
        LEFT JOIN meetings m ON m.meeting_id=su.meeting_id AND m.starts_at >= %s AND m.starts_at < %s
        WHERE cr.start_date IS NULL OR cr.start_date < %s
        GROUP BY c.class_code,co.course_code,co.course_name,cr.course_run_id
        ORDER BY c.class_code,co.course_name,cr.course_run_id
    """, (review_month, next_month, next_month))
    participation = fetch_all(pool, """
        SELECT c.class_code,co.course_code,co.course_name,re.run_enrollment_id,e.emp_code,e.full_name,
               cr.attendance_threshold_ratio_snapshot AS attendance_threshold,
               count(DISTINCT su.sequence_in_run) AS applicable_sessions,
               count(DISTINCT su.sequence_in_run) FILTER (
                   WHERE a.effective_status='Present' OR makeup_meeting.meeting_id IS NOT NULL
               ) AS present_sessions,
               round(count(DISTINCT su.sequence_in_run) FILTER (
                         WHERE a.effective_status='Present' OR makeup_meeting.meeting_id IS NOT NULL
                     )::numeric /
                     NULLIF(count(DISTINCT su.sequence_in_run),0),4) AS attendance_ratio
        FROM run_enrollments re
        JOIN session_units su
          ON su.course_run_id=re.course_run_id
         AND su.sequence_in_run>=re.start_session_number
         AND su.unit_type<>'makeup'
        JOIN meetings m ON m.meeting_id=su.meeting_id AND m.status='completed'
        LEFT JOIN attendance a
          ON a.run_enrollment_id=re.run_enrollment_id
         AND a.session_unit_id=su.session_unit_id
        LEFT JOIN attendance makeup
          ON makeup.makeup_for_attendance_id=a.attendance_id
         AND makeup.is_makeup
         AND makeup.effective_status='Present'
        LEFT JOIN session_units makeup_unit ON makeup_unit.session_unit_id=makeup.session_unit_id
        LEFT JOIN meetings makeup_meeting
          ON makeup_meeting.meeting_id=makeup_unit.meeting_id
         AND makeup_meeting.status='completed'
        JOIN employees e ON e.employee_id=re.employee_id JOIN course_runs cr ON cr.course_run_id=re.course_run_id
        JOIN cohorts c ON c.cohort_id=cr.cohort_id JOIN courses co ON co.course_id=cr.course_id
        WHERE m.starts_at >= %s AND m.starts_at < %s
        GROUP BY c.class_code,co.course_code,co.course_name,re.run_enrollment_id,e.emp_code,e.full_name,cr.attendance_threshold_ratio_snapshot
        ORDER BY c.class_code,co.course_name,e.full_name
    """, (review_month, next_month))
    progress = fetch_all(pool, """
        WITH ordered AS (
            SELECT ev.evaluation_version_id,e.run_enrollment_id,re.employee_id,cr.course_id,co.course_name,
                   ev.final_level_id,ev.created_at,
                   lag(ev.final_level_id) OVER (PARTITION BY re.employee_id ORDER BY ev.created_at,ev.evaluation_version_id) AS prior_level_id,
                   row_number() OVER (PARTITION BY re.employee_id ORDER BY ev.created_at DESC,ev.evaluation_version_id DESC) AS latest_rank
            FROM evaluation_versions ev JOIN evaluations e ON e.evaluation_id=ev.evaluation_id
            JOIN run_enrollments re ON re.run_enrollment_id=e.run_enrollment_id JOIN course_runs cr ON cr.course_run_id=re.course_run_id
            JOIN courses co ON co.course_id=cr.course_id WHERE ev.final_level_id IS NOT NULL
        )
        SELECT o.course_id,o.course_name,emp.emp_code,emp.full_name,current_level.level_name AS latest_level,
               prior_level.level_name AS prior_level,(current_level.numeric_value > prior_level.numeric_value) AS improved
        FROM ordered o JOIN employees emp ON emp.employee_id=o.employee_id
        JOIN levels current_level ON current_level.level_id=o.final_level_id
        LEFT JOIN levels prior_level ON prior_level.level_id=o.prior_level_id
        WHERE o.latest_rank=1 AND o.created_at >= %s AND o.created_at < %s
        ORDER BY o.course_name,emp.full_name
    """, (review_month, next_month))
    new_courses = fetch_all(pool, """SELECT course_code,course_name,expected_units,attendance_threshold_ratio,created_at
                                     FROM courses WHERE created_at >= %s AND created_at < %s ORDER BY created_at,course_name""", (review_month, next_month))
    action_versions = fetch_all(pool, """SELECT summary.version_number,summary.highlights,summary.risks,
                                                summary.next_month_priorities,summary.created_at,
                                                actor.username AS created_by_username
                                         FROM monthly_review_action_summary_versions summary
                                         JOIN app_users actor ON actor.user_id=summary.created_by_user_id
                                         WHERE review_month=%s ORDER BY version_number DESC LIMIT 1""", (review_month,))
    level_distribution: dict[tuple[str, str], int] = {}
    for row in progress:
        key = (row["course_name"], row["latest_level"])
        level_distribution[key] = level_distribution.get(key, 0) + 1
    course_participation: dict[tuple[str, str], dict[str, object]] = {}
    class_participation: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in participation:
        course_key = (row["course_code"], row["course_name"])
        class_key = (row["class_code"], row["course_code"], row["course_name"])
        for target, key, labels in (
            (course_participation, course_key, {"course_code": row["course_code"], "course_name": row["course_name"]}),
            (class_participation, class_key, {"class_code": row["class_code"], "course_code": row["course_code"], "course_name": row["course_name"]}),
        ):
            bucket = target.setdefault(
                key,
                {**labels, "applicable_sessions": 0, "present_sessions": 0, "learner_count": 0, "low_attendance_count": 0},
            )
            bucket["applicable_sessions"] += row["applicable_sessions"] or 0
            bucket["present_sessions"] += row["present_sessions"] or 0
            bucket["learner_count"] += 1
            if row["attendance_ratio"] is not None and row["attendance_ratio"] < row["attendance_threshold"]:
                bucket["low_attendance_count"] += 1
    for rows in (course_participation.values(), class_participation.values()):
        for row in rows:
            row["attendance_ratio"] = (
                round(row["present_sessions"] / row["applicable_sessions"], 4)
                if row["applicable_sessions"]
                else None
            )
    return {"program": program, "participation": participation, "progress": progress, "new_courses": new_courses,
            "course_participation": sorted(course_participation.values(), key=lambda row: (row["course_name"], row["course_code"])),
            "class_participation": sorted(class_participation.values(), key=lambda row: (row["class_code"], row["course_name"])),
            "level_distribution": [{"course_name": course, "latest_level": level, "learner_count": count}
                                   for (course, level), count in sorted(level_distribution.items())],
            "action_summary": action_versions[0] if action_versions else None}


def monthly_review_summary(data: dict[str, object]) -> dict[str, object]:
    program = data["program"]
    participation = data["participation"]
    progress = data["progress"]
    active = sum(row["active_participants"] for row in program)
    repeated = sum(row["repeated_participants"] for row in program)
    planned = sum(row["planned_sessions"] for row in program)
    delivered = sum(row["delivered_sessions"] for row in program)
    total = sum(row["applicable_sessions"] for row in participation)
    present = sum(row["present_sessions"] for row in participation)
    low = [row for row in participation if row["attendance_ratio"] is not None and row["attendance_ratio"] < row["attendance_threshold"]]
    improved = [row for row in progress if row["improved"] is True]
    return {"active": active, "repeated": repeated, "planned": planned, "delivered": delivered,
            "variance": delivered-planned, "attendance_ratio": (present / total) if total else None,
            "low_count": len(low), "improved_count": len(improved), "tested_count": len(progress),
            "delivery_rate": (delivered / planned) if planned else None,
            "low_rate": (len(low) / len(participation)) if participation else None,
            "improved_rate": (len(improved) / len(progress)) if progress else None,
            "new_course_count": len(data["new_courses"])}


def proposed_monthly_actions(summary: dict[str, object]) -> dict[str, str]:
    highlights = f"{summary['delivered']} delivered logical session(s); {summary['active']} active participant(s)."
    risks = "No attendance risk identified." if not summary["low_count"] else f"{summary['low_count']} active learner(s) are below their run attendance threshold."
    priorities = "Review upcoming schedule and learner follow-ups."
    if summary["low_count"]:
        priorities = "Contact low-attendance learners and confirm make-up or recovery actions."
    return {"highlights": highlights, "risks": risks, "next_month_priorities": priorities}


def monthly_review_xlsx(review_month: date, data: dict[str, object], action_summary: dict[str, str]) -> bytes:
    """Create a portable, inspectable Excel export from the displayed canonical detail tables."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = [("Program status", data["program"]), ("Participation", data["participation"]),
              ("Course participation", data["course_participation"]), ("Class participation", data["class_participation"]),
              ("Learning progress", data["progress"]), ("Level distribution", data["level_distribution"]),
              ("New courses", data["new_courses"])]
    for title, rows in sheets:
        worksheet = workbook.create_sheet(title)
        worksheet.append([f"Monthly review: {review_month.isoformat()}"])
        headers = list(rows[0].keys()) if rows else []
        if headers:
            worksheet.append(headers)
            for cell in worksheet[2]: cell.font = Font(bold=True)
            for row in rows: worksheet.append(list(row.values()))
            worksheet.freeze_panes = "A3"
            for column in worksheet.columns:
                worksheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 36)
    action = workbook.create_sheet("Action summary")
    action.append([f"Monthly review: {review_month.isoformat()}"])
    for heading, value in (("Highlights", action_summary["highlights"]), ("Risks", action_summary["risks"]), ("Next-month priorities", action_summary["next_month_priorities"])):
        action.append([heading]); action.append([value]); action[action.max_row - 1][0].font = Font(bold=True)
    output = BytesIO(); workbook.save(output); return output.getvalue()
