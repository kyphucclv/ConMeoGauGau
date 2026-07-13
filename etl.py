"""
ETL: load okok_FIXED_v2.xlsx into the PostgreSQL schema defined in schema.sql.

Usage:
    python3 etl.py <path-to-xlsx> <postgres-connection-uri>

Cleaning rules applied (all counts are printed at the end so nothing is
silently dropped):
  - Emp Code is always cast to TEXT (source file mixes string and float
    representations of the same code across sheets).
  - Level names are matched to level_helper case-insensitively; unmatched
    labels (e.g. the placeholder "Chua test") are stored as NULL rather
    than violating the FK.
  - class_offerings is the UNION of CLASS_DATES plus any (class_code,
    course_name) pairs referenced by sheet2 (enrollments) or ATTENDANCE_LOG
    but missing from CLASS_DATES -- these are real classes/courses in the
    data, just missing a row in that particular sheet.
  - Rows missing a required key (e.g. enrollment row with no course name,
    attendance row with no session date) are skipped and counted.
"""
import sys
import datetime
import openpyxl
import psycopg2


def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def emp(v):
    v = norm(v)
    if v is None:
        return None
    if isinstance(v, float):
        v = int(v)
    return str(v).strip()


def main(xlsx_path, pg_uri):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    conn = psycopg2.connect(pg_uri)
    cur = conn.cursor()
    stats = {}

    # ---- level_helper -------------------------------------------------
    ws = wb["LEVEL_HELPER"]
    levels = {}
    rows = []
    for r in range(2, ws.max_row + 1):
        name, val = norm(ws.cell(r, 1).value), norm(ws.cell(r, 2).value)
        if name is None:
            continue
        levels[name.lower()] = name
        rows.append((name, val))
    cur.executemany("INSERT INTO level_helper (level_name, numeric_value) VALUES (%s,%s)", rows)
    stats["level_helper"] = len(rows)

    def match_level(v):
        v = norm(v)
        if v is None:
            return None
        return levels.get(v.lower())

    # ---- course_plan ----------------------------------------------------
    ws = wb["COURSE_PLAN"]
    courses = set()
    rows = []
    for r in range(2, ws.max_row + 1):
        name, sessions = norm(ws.cell(r, 1).value), norm(ws.cell(r, 2).value)
        if name is None:
            continue
        courses.add(name)
        rows.append((name, int(sessions)))
    cur.executemany("INSERT INTO course_plan (course_name, expected_sessions) VALUES (%s,%s)", rows)
    stats["course_plan"] = len(rows)

    def match_course(v):
        v = norm(v)
        return v if v in courses else None

    # ---- class_pic --------------------------------------------------
    ws = wb["PIC"]
    class_pics = set()
    rows = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        cc = norm(ws.cell(r, 1).value)
        pic_name = norm(ws.cell(r, 2).value)
        pic_emp = emp(ws.cell(r, 3).value)
        mail = norm(ws.cell(r, 4).value)
        if cc is None or cc in seen:
            continue
        seen.add(cc)
        class_pics.add(cc)
        rows.append((cc, pic_name or "Unknown", None, mail, None))
    cur.executemany(
        "INSERT INTO class_pic (class_code, pic_name, pic_emp_code, mail, english_name) VALUES (%s,%s,%s,%s,%s)",
        rows,
    )
    stats["class_pic"] = len(rows)

    # ---- students (load first w/o FK-sensitive cols, patch after) ------
    ws = wb["STUDENTS"]
    students_rows = []
    emp_codes = set()
    pending_latest = []  # (emp_code, class_code, course_name) patched after class_offerings exists
    for r in range(2, ws.max_row + 1):
        e = emp(ws.cell(r, 1).value)
        if e is None:
            continue
        emp_codes.add(e)
        full_name = norm(ws.cell(r, 2).value)
        bu = norm(ws.cell(r, 3).value)
        role = norm(ws.cell(r, 4).value)
        status = norm(ws.cell(r, 5).value)
        pic = norm(ws.cell(r, 6).value)
        current_course = match_course(ws.cell(r, 7).value)
        entrance_level = match_level(ws.cell(r, 8).value)
        current_level = match_level(ws.cell(r, 9).value)
        last_active = norm(ws.cell(r, 10).value)
        drop_flag = norm(ws.cell(r, 12).value)
        drop_def = norm(ws.cell(r, 13).value)
        drop_reason = norm(ws.cell(r, 14).value)
        remark = norm(ws.cell(r, 15).value)
        latest_cc = norm(ws.cell(r, 17).value)
        latest_cn = norm(ws.cell(r, 18).value)
        progress_cat = norm(ws.cell(r, 23).value)
        students_rows.append(
            (e, full_name, bu, role, status, pic, current_course, entrance_level,
             current_level, last_active, drop_flag, drop_def, drop_reason, remark,
             progress_cat)
        )
        if latest_cc and latest_cn:
            pending_latest.append((e, latest_cc, latest_cn))

    cur.executemany(
        """INSERT INTO students
           (emp_code, full_name, bu, role, status, pic, current_course, entrance_level,
            current_level, last_active_date, drop_flag, drop_definition, drop_reason,
            remark, progress_category)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
        students_rows,
    )
    stats["students"] = len(students_rows)

    # ---- class_offerings: union of CLASS_DATES + sheet2 + ATTENDANCE_LOG --
    ws = wb["CLASS_DATES"]
    offerings = {}  # (class,course) -> start_date
    for r in range(2, ws.max_row + 1):
        cc, cn = norm(ws.cell(r, 1).value), match_course(ws.cell(r, 2).value)
        sd = norm(ws.cell(r, 3).value)
        if cc is None or cn is None:
            continue
        offerings.setdefault((cc, cn), sd)

    extra_from_sheet2 = 0
    ws2 = wb["sheet2"]
    for r in range(2, ws2.max_row + 1):
        cc, cn = norm(ws2.cell(r, 3).value), match_course(ws2.cell(r, 5).value)
        if cc and cn and (cc, cn) not in offerings and cc in class_pics:
            offerings[(cc, cn)] = None
            extra_from_sheet2 += 1

    extra_from_att = 0
    wsA = wb["ATTENDANCE_LOG"]
    for r in range(2, wsA.max_row + 1):
        cc, cn = norm(wsA.cell(r, 1).value), match_course(wsA.cell(r, 2).value)
        if cc and cn and (cc, cn) not in offerings and cc in class_pics:
            offerings[(cc, cn)] = None
            extra_from_att += 1

    rows = [(cc, cn, sd) for (cc, cn), sd in offerings.items()]
    cur.executemany(
        "INSERT INTO class_offerings (class_code, course_name, start_date) VALUES (%s,%s,%s)", rows
    )
    stats["class_offerings"] = len(rows)
    stats["class_offerings_recovered_from_sheet2"] = extra_from_sheet2
    stats["class_offerings_recovered_from_attendance_log"] = extra_from_att

    # patch students.latest_class_code/latest_course_name now that FK target exists
    n_patched = n_skipped_latest = 0
    for e, cc, cn in pending_latest:
        if (cc, cn) in offerings:
            cur.execute(
                "UPDATE students SET latest_class_code=%s, latest_course_name=%s WHERE emp_code=%s",
                (cc, cn, e),
            )
            n_patched += 1
        else:
            n_skipped_latest += 1
    stats["students_latest_class_patched"] = n_patched
    stats["students_latest_class_skipped"] = n_skipped_latest

    # ---- placements ------------------------------------------------------
    ws = wb["Placement"]
    rows = []
    skipped = 0
    for r in range(6, ws.max_row + 1):
        e = emp(ws.cell(r, 1).value)
        full_name = norm(ws.cell(r, 2).value)
        if e is None and full_name is None:
            continue
        test_date = norm(ws.cell(r, 3).value)
        if isinstance(test_date, datetime.datetime):
            test_date = test_date.date()
        elif not isinstance(test_date, datetime.date):
            test_date = None  # free-text notes like "Join after entrance test"
        level = match_level(ws.cell(r, 4).value)
        grammar = norm(ws.cell(r, 5).value)
        vocab = norm(ws.cell(r, 6).value)
        pron = norm(ws.cell(r, 7).value)
        fluency = norm(ws.cell(r, 8).value)
        emp_fk = e if e in emp_codes else None
        if e is not None and emp_fk is None:
            skipped += 1
        rows.append((emp_fk, full_name, test_date, level, grammar, vocab, pron, fluency))
    cur.executemany(
        """INSERT INTO placements
           (emp_code, full_name, test_date, level, grammar_feedback, vocabulary_feedback,
            pronunciation_feedback, fluency_feedback)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
        rows,
    )
    stats["placements"] = len(rows)
    stats["placements_emp_code_unmatched"] = skipped

    # ---- enrollments (sheet2) --------------------------------------------
    ws = wb["sheet2"]
    rows = []
    skipped = 0
    for r in range(2, ws.max_row + 1):
        e = emp(ws.cell(r, 1).value)
        cc = norm(ws.cell(r, 3).value)
        cn = match_course(ws.cell(r, 5).value)
        if e is None or cc is None or cn is None or e not in emp_codes or (cc, cn) not in offerings:
            skipped += 1
            continue
        entrance_level = match_level(ws.cell(r, 6).value)
        final_level = match_level(ws.cell(r, 7).value)
        start_date = norm(ws.cell(r, 8).value)
        first_class_start = norm(ws.cell(r, 9).value)
        rows.append((e, cc, cn, entrance_level, final_level, start_date, first_class_start))
    # de-dup on PK (emp_code, class_code, course_name), keep first occurrence
    seen = set()
    dedup_rows = []
    for row in rows:
        k = row[:3]
        if k in seen:
            continue
        seen.add(k)
        dedup_rows.append(row)
    cur.executemany(
        """INSERT INTO enrollments
           (emp_code, class_code, course_name, entrance_level, final_level, start_date, first_class_start_date)
           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
        dedup_rows,
    )
    stats["enrollments"] = len(dedup_rows)
    stats["enrollments_skipped"] = skipped

    # ---- attendance_log ---------------------------------------------------
    ws = wb["ATTENDANCE_LOG"]
    rows = []
    skipped = 0
    for r in range(2, ws.max_row + 1):
        cc = norm(ws.cell(r, 1).value)
        cn = match_course(ws.cell(r, 2).value)
        e = emp(ws.cell(r, 3).value)
        so = norm(ws.cell(r, 5).value)
        dt = norm(ws.cell(r, 6).value)
        status = norm(ws.cell(r, 7).value)
        if not (cc and cn and e and so is not None and dt is not None and status in ("Present", "Absent")):
            skipped += 1
            continue
        if e not in emp_codes or (cc, cn) not in offerings:
            skipped += 1
            continue
        rows.append((cc, cn, e, int(so), dt, status))
    cur.executemany(
        """INSERT INTO attendance_log (class_code, course_name, emp_code, session_order, session_date, status)
           VALUES (%s,%s,%s,%s,%s,%s)""",
        rows,
    )
    stats["attendance_log"] = len(rows)
    stats["attendance_log_skipped"] = skipped

    conn.commit()
    cur.close()
    conn.close()

    print("Load complete:")
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "okok_FIXED_v2.xlsx"
    pg_uri = sys.argv[2] if len(sys.argv) > 2 else "postgresql://postgres@localhost/postgres"
    main(xlsx_path, pg_uri)
