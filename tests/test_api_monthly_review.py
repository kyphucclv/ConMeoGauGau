from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
from threading import Barrier

from openpyxl import load_workbook

from tests.test_api_auth import ORIGIN, client_for


def _login(client, username="pytest_editor", password="editor-pass"):
    response = client.post(
        "/api/auth/login",
        headers={"Origin": ORIGIN},
        json={"username": username, "password": password},
    )
    assert response.status_code == 200
    return response.json()


def test_editor_reviews_one_normalized_month_through_the_http_interface(
    database_url, factory
):
    _, course_run_id = factory.cohort_run()
    factory.onboard(course_run_id, full_name="Monthly Review Alpha")
    factory.meeting_unit(course_run_id, 1, status="completed")
    pool, client = client_for(database_url)
    try:
        with client:
            _login(client)
            response = client.get("/api/monthly-review?month=2026-08")

        assert response.status_code == 200
        body = response.json()
        assert body["review_month"] == "2026-08-01"
        assert body["summary"]["active"] >= 1
        assert body["summary"]["delivered"] >= 1
        assert body["program"]
        assert body["action_summary"] is None
        assert body["proposed_action_summary"]["highlights"]
    finally:
        pool.closeall()


def test_editor_saves_immutable_monthly_conclusions_with_named_attribution(
    database_url, seed_ids
):
    pool, client = client_for(database_url)
    try:
        with client:
            auth = _login(client)
            first = client.post(
                "/api/monthly-review/action-summary",
                headers={"X-CSRF-Token": auth["csrf_token"]},
                json={
                    "month": "2026-09",
                    "highlights": "First conclusion",
                    "risks": "First risk",
                    "next_month_priorities": "First priority",
                },
            )
            second = client.post(
                "/api/monthly-review/action-summary",
                headers={"X-CSRF-Token": auth["csrf_token"]},
                json={
                    "month": "2026-09",
                    "highlights": "Corrected conclusion",
                    "risks": "Corrected risk",
                    "next_month_priorities": "Corrected priority",
                },
            )
            reviewed = client.get("/api/monthly-review?month=2026-09")

        assert first.status_code == 200
        assert first.json()["version_number"] == 1
        assert second.status_code == 200
        assert second.json()["version_number"] == 2
        assert reviewed.json()["action_summary"]["version_number"] == 2
        assert reviewed.json()["action_summary"]["highlights"] == "Corrected conclusion"
        assert reviewed.json()["action_summary"]["created_by_username"] == "pytest_editor"
    finally:
        pool.closeall()


def test_editor_exports_the_displayed_month_as_a_private_named_workbook(
    database_url, seed_ids
):
    pool, client = client_for(database_url)
    try:
        with client:
            auth = _login(client)
            saved = client.post(
                "/api/monthly-review/action-summary",
                headers={"X-CSRF-Token": auth["csrf_token"]},
                json={
                    "month": "2026-10",
                    "highlights": "Workbook highlight",
                    "risks": "Workbook risk",
                    "next_month_priorities": "Workbook priority",
                },
            )
            assert saved.status_code == 200
            exported = client.get("/api/monthly-review/export?month=2026-10")

        assert exported.status_code == 200
        assert exported.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert exported.headers["cache-control"] == "private, no-store"
        assert exported.headers["content-disposition"] == (
            'attachment; filename="english-class-monthly-review-2026-10-01.xlsx"'
        )
        workbook = load_workbook(BytesIO(exported.content), read_only=True, data_only=False)
        assert workbook.sheetnames == [
            "Program status",
            "Participation",
            "Course participation",
            "Class participation",
            "Learning progress",
            "Level distribution",
            "New courses",
            "Action summary",
        ]
        action = workbook["Action summary"]
        assert action["A1"].value == "Monthly review: 2026-10-01"
        assert action["A3"].value == "Workbook highlight"
        assert action["A5"].value == "Workbook risk"
        assert action["A7"].value == "Workbook priority"
    finally:
        pool.closeall()


def test_monthly_review_rejects_viewer_bad_month_csrf_and_forged_fields(
    database_url, seed_ids
):
    pool, client = client_for(database_url)
    try:
        with client:
            viewer = _login(client, "pytest_viewer", "viewer-pass")
            assert client.get("/api/monthly-review?month=2026-08").status_code == 403
            assert client.get("/api/monthly-review/export?month=2026-08").status_code == 403
            client.post(
                "/api/auth/logout",
                headers={"X-CSRF-Token": viewer["csrf_token"]},
            )
            editor = _login(client)
            invalid_month = client.get("/api/monthly-review?month=2026-13")
            missing_csrf = client.post(
                "/api/monthly-review/action-summary",
                json={"month": "2026-08", "highlights": "", "risks": "", "next_month_priorities": ""},
            )
            forged = client.post(
                "/api/monthly-review/action-summary",
                headers={"X-CSRF-Token": editor["csrf_token"]},
                json={
                    "month": "2026-08",
                    "highlights": "Safe",
                    "risks": "Safe",
                    "next_month_priorities": "Safe",
                    "version_number": 99,
                    "created_by_username": "forged",
                },
            )

        assert invalid_month.status_code == 422
        assert invalid_month.json()["code"] == "invalid_input"
        assert missing_csrf.status_code == 403
        assert missing_csrf.json()["code"] == "csrf_rejected"
        assert forged.status_code == 422
        assert forged.json()["code"] == "invalid_input"
    finally:
        pool.closeall()


def test_concurrent_monthly_conclusions_receive_distinct_versions(database_url, seed_ids):
    barrier = Barrier(2)

    def submit(label: str):
        pool, client = client_for(database_url)
        try:
            with client:
                auth = _login(client)
                barrier.wait()
                return client.post(
                    "/api/monthly-review/action-summary",
                    headers={"X-CSRF-Token": auth["csrf_token"]},
                    json={
                        "month": "2027-01",
                        "highlights": label,
                        "risks": f"{label} risk",
                        "next_month_priorities": f"{label} priority",
                    },
                )
        finally:
            pool.closeall()

    with ThreadPoolExecutor(max_workers=2) as executor:
        responses = list(executor.map(submit, ["Concurrent A", "Concurrent B"]))

    assert [response.status_code for response in responses] == [200, 200]
    assert sorted(response.json()["version_number"] for response in responses) == [1, 2]
