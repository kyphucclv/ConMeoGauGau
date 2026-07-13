"""Stage workbook rows and produce an auditable profile report."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


CORE_SHEETS = {
    "STUDENTS",
    "sheet2",
    "ATTENDANCE_LOG",
    "Placement",
    "PIC",
    "LEVEL_HELPER",
    "CLASS_DATES",
    "COURSE_PLAN",
}

FIELD_ALIASES = {
    "emp_code": {
        "STUDENTS": ["Emp Code"],
        "sheet2": ["Emp Code"],
        "ATTENDANCE_LOG": ["Emp Code"],
        "Placement": ["Emp. Code"],
        "PIC": ["EMP Code"],
    },
    "class_code": {
        "PIC": ["Class Code"],
        "CLASS_DATES": ["Class Code"],
        "sheet2": ["Class Code"],
        "ATTENDANCE_LOG": ["Class Code"],
    },
    "course_name": {
        "COURSE_PLAN": ["Course Name"],
        "CLASS_DATES": ["Course Name"],
        "sheet2": ["Course Name"],
        "ATTENDANCE_LOG": ["Course Name"],
        "STUDENTS": ["Current Course", "Latest Course Name"],
    },
    "level": {
        "LEVEL_HELPER": ["Level Name"],
        "STUDENTS": ["Entrance Level", "Current Level"],
        "sheet2": ["Entrance Level", "Final Level"],
        "Placement": ["1st session:"],
    },
}


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "text") and hasattr(value, "ref"):
        return {
            "formula_type": type(value).__name__,
            "ref": getattr(value, "ref", None),
            "text": getattr(value, "text", None),
        }
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def normalized_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.strip().split())
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def infer_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float, Decimal)):
        return "number"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, str):
        if value.startswith("="):
            return "formula"
        if value.startswith("#"):
            return "error"
        return "text"
    if hasattr(value, "text") and hasattr(value, "ref"):
        return "formula"
    return type(value).__name__


def stable_hash(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, default=json_value)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def workbook_checksum(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def unique_headers(values: list[Any], max_columns: int) -> list[str]:
    seen: Counter[str] = Counter()
    headers: list[str] = []
    for idx in range(max_columns):
        raw = values[idx] if idx < len(values) else None
        base = normalized_text(raw) or f"column_{idx + 1}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def profile_workbook(path: Path) -> dict[str, Any]:
    checksum = workbook_checksum(path)
    wb = load_workbook(path, read_only=True, data_only=False)
    workbook_profile: dict[str, Any] = {
        "source_name": path.name,
        "source_checksum": checksum,
        "file_size_bytes": path.stat().st_size,
        "sheets": [],
    }

    for ws in wb.worksheets:
        rows_payload: list[dict[str, Any]] = []
        physical_rows = 0
        meaningful_rows = 0
        max_columns = 0
        formula_cells = 0
        error_cells = 0
        header_values: list[Any] | None = None
        headers: list[str] = []
        field_values: dict[int, list[Any]] = defaultdict(list)
        field_type_counts: dict[int, Counter[str]] = defaultdict(Counter)
        malformed_examples: dict[int, list[dict[str, Any]]] = defaultdict(list)

        for row in ws.iter_rows():
            physical_rows += 1
            cells = list(row)
            values = [cell.value for cell in cells]
            max_columns = max(max_columns, len(values))
            if not any(value is not None for value in values):
                continue

            meaningful_rows += 1
            if header_values is None:
                header_values = values
                headers = unique_headers(header_values, max_columns)

            if len(headers) < len(values):
                headers = unique_headers(header_values or [], len(values))

            source_row_number = physical_rows
            raw_cells: list[dict[str, Any]] = []
            row_by_header: dict[str, Any] = {}
            for idx, cell in enumerate(cells, start=1):
                value = json_value(cell.value)
                header = headers[idx - 1] if idx - 1 < len(headers) else f"column_{idx}"
                data_type = infer_type(cell.value)
                if data_type == "formula":
                    formula_cells += 1
                if data_type == "error":
                    error_cells += 1
                    if len(malformed_examples[idx]) < 10:
                        malformed_examples[idx].append(
                            {"source_row_number": source_row_number, "value": value}
                        )

                raw_cells.append(
                    {
                        "column_index": idx,
                        "coordinate": f"{get_column_letter(idx)}{source_row_number}",
                        "header": header,
                        "value": value,
                        "data_type": data_type,
                    }
                )
                row_by_header[header] = value

                if meaningful_rows > 1:
                    field_values[idx].append(value)
                    field_type_counts[idx][data_type] += 1

            row_hash = stable_hash(
                {
                    "sheet_name": ws.title,
                    "source_row_number": source_row_number,
                    "cells": raw_cells,
                }
            )
            rows_payload.append(
                {
                    "source_row_number": source_row_number,
                    "row_hash": row_hash,
                    "raw_payload": {
                        "sheet_name": ws.title,
                        "source_row_number": source_row_number,
                        "headers": headers,
                        "values_by_header": row_by_header,
                        "cells": raw_cells,
                    },
                }
            )

        row_hash_counts = Counter(row["row_hash"] for row in rows_payload)
        fields: list[dict[str, Any]] = []
        data_row_count = max(meaningful_rows - 1, 0)
        for idx, header in enumerate(headers, start=1):
            values = field_values.get(idx, [])
            non_null_values = [v for v in values if v is not None and v != ""]
            value_counts = Counter(json.dumps(v, ensure_ascii=False, default=json_value) for v in non_null_values)
            fields.append(
                {
                    "field_name": header,
                    "column_index": idx,
                    "non_null_count": len(non_null_values),
                    "null_count": max(data_row_count - len(non_null_values), 0),
                    "distinct_count": len(value_counts),
                    "duplicate_non_null_count": sum(count - 1 for count in value_counts.values() if count > 1),
                    "inferred_types": dict(field_type_counts.get(idx, Counter())),
                    "top_values": [
                        {"value": json.loads(value), "count": count}
                        for value, count in value_counts.most_common(10)
                    ],
                    "malformed_examples": malformed_examples.get(idx, []),
                }
            )

        workbook_profile["sheets"].append(
            {
                "sheet_name": ws.title,
                "is_core_sheet": ws.title in CORE_SHEETS,
                "physical_rows": physical_rows,
                "meaningful_rows": meaningful_rows,
                "max_columns": max_columns,
                "formula_cells": formula_cells,
                "error_cells": error_cells,
                "duplicate_row_hashes": sum(1 for count in row_hash_counts.values() if count > 1),
                "fields": fields,
                "rows": rows_payload,
            }
        )

    workbook_profile["cross_sheet_coverage"] = cross_sheet_coverage(workbook_profile)
    return workbook_profile


def values_for_alias(sheet: dict[str, Any], field_names: list[str]) -> set[str]:
    values: set[str] = set()
    for row in sheet["rows"]:
        if row["source_row_number"] == 1:
            continue
        by_header = row["raw_payload"]["values_by_header"]
        for field_name in field_names:
            value = normalized_text(by_header.get(field_name))
            if value:
                values.add(value)
    return values


def cross_sheet_coverage(profile: dict[str, Any]) -> dict[str, Any]:
    sheets = {sheet["sheet_name"]: sheet for sheet in profile["sheets"]}
    coverage: dict[str, Any] = {}
    reference_sheet = {
        "emp_code": "STUDENTS",
        "class_code": "PIC",
        "course_name": "COURSE_PLAN",
        "level": "LEVEL_HELPER",
    }

    for key_name, aliases in FIELD_ALIASES.items():
        sheet_sets: dict[str, set[str]] = {}
        for sheet_name, field_names in aliases.items():
            if sheet_name in sheets:
                sheet_sets[sheet_name] = values_for_alias(sheets[sheet_name], field_names)

        ref_name = reference_sheet[key_name]
        reference_values = sheet_sets.get(ref_name, set())
        coverage[key_name] = {
            "reference_sheet": ref_name,
            "reference_distinct": len(reference_values),
            "sheets": {},
        }
        for sheet_name, values in sheet_sets.items():
            missing_from_reference = sorted(values - reference_values)
            missing_reference_values = sorted(reference_values - values)
            coverage[key_name]["sheets"][sheet_name] = {
                "distinct": len(values),
                "missing_from_reference_count": len(missing_from_reference),
                "missing_from_reference_examples": missing_from_reference[:20],
                "reference_values_absent_count": len(missing_reference_values),
                "reference_values_absent_examples": missing_reference_values[:20],
            }

    return coverage


def load_profile(conn, profile: dict[str, Any]) -> dict[str, int]:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO source_workbooks (source_name, source_checksum, file_size_bytes, workbook_metadata)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (source_name, source_checksum)
            DO UPDATE SET workbook_metadata = EXCLUDED.workbook_metadata
            RETURNING source_workbook_id
            """,
            (
                profile["source_name"],
                profile["source_checksum"],
                profile["file_size_bytes"],
                psycopg2.extras.Json({"sheet_count": len(profile["sheets"])}),
            ),
        )
        source_workbook_id = cur.fetchone()[0]

        cur.execute(
            """
            INSERT INTO import_batches (source_name, source_checksum, status, stats, completed_at)
            VALUES (%s, %s, 'completed', %s, NOW())
            ON CONFLICT (source_name, source_checksum)
            DO UPDATE SET stats = EXCLUDED.stats, completed_at = NOW(), status = 'completed'
            RETURNING import_batch_id
            """,
            (
                profile["source_name"],
                profile["source_checksum"],
                psycopg2.extras.Json({"profiled_by": "scripts/stage_workbook.py"}),
            ),
        )
        import_batch_id = cur.fetchone()[0]

        inserted_rows = 0
        for sheet in profile["sheets"]:
            cur.execute(
                """
                INSERT INTO workbook_sheet_profiles (
                    source_workbook_id, sheet_name, physical_rows, meaningful_rows,
                    max_columns, formula_cells, error_cells, duplicate_row_hashes, profile
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (source_workbook_id, sheet_name)
                DO UPDATE SET
                    physical_rows = EXCLUDED.physical_rows,
                    meaningful_rows = EXCLUDED.meaningful_rows,
                    max_columns = EXCLUDED.max_columns,
                    formula_cells = EXCLUDED.formula_cells,
                    error_cells = EXCLUDED.error_cells,
                    duplicate_row_hashes = EXCLUDED.duplicate_row_hashes,
                    profile = EXCLUDED.profile,
                    profiled_at = NOW()
                RETURNING sheet_profile_id
                """,
                (
                    source_workbook_id,
                    sheet["sheet_name"],
                    sheet["physical_rows"],
                    sheet["meaningful_rows"],
                    sheet["max_columns"],
                    sheet["formula_cells"],
                    sheet["error_cells"],
                    sheet["duplicate_row_hashes"],
                    psycopg2.extras.Json({"is_core_sheet": sheet["is_core_sheet"]}),
                ),
            )
            sheet_profile_id = cur.fetchone()[0]

            cur.execute("DELETE FROM workbook_field_profiles WHERE sheet_profile_id = %s", (sheet_profile_id,))
            for field in sheet["fields"]:
                cur.execute(
                    """
                    INSERT INTO workbook_field_profiles (
                        sheet_profile_id, field_name, column_index, non_null_count,
                        null_count, distinct_count, duplicate_non_null_count,
                        inferred_types, top_values, malformed_examples
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        sheet_profile_id,
                        field["field_name"],
                        field["column_index"],
                        field["non_null_count"],
                        field["null_count"],
                        field["distinct_count"],
                        field["duplicate_non_null_count"],
                        psycopg2.extras.Json(field["inferred_types"]),
                        psycopg2.extras.Json(field["top_values"]),
                        psycopg2.extras.Json(field["malformed_examples"]),
                    ),
                )

            for row in sheet["rows"]:
                cur.execute(
                    """
                    INSERT INTO raw_workbook_rows (
                        import_batch_id, source_workbook_id, source_name, source_checksum,
                        sheet_name, source_row_number, row_hash, raw_payload
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (import_batch_id, sheet_name, source_row_number) DO NOTHING
                    """,
                    (
                        import_batch_id,
                        source_workbook_id,
                        profile["source_name"],
                        profile["source_checksum"],
                        sheet["sheet_name"],
                        row["source_row_number"],
                        row["row_hash"],
                        psycopg2.extras.Json(row["raw_payload"]),
                    ),
                )
                inserted_rows += cur.rowcount

    return {"source_workbook_id": source_workbook_id, "import_batch_id": import_batch_id, "inserted_rows": inserted_rows}


def profile_for_json(profile: dict[str, Any]) -> dict[str, Any]:
    compact = dict(profile)
    compact["sheets"] = []
    for sheet in profile["sheets"]:
        sheet_copy = {k: v for k, v in sheet.items() if k != "rows"}
        sheet_copy["staged_rows"] = len(sheet["rows"])
        compact["sheets"].append(sheet_copy)
    return compact


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--database-url")
    parser.add_argument("--profile-output", type=Path)
    args = parser.parse_args()

    profile = profile_workbook(args.workbook)
    if args.profile_output:
        args.profile_output.parent.mkdir(parents=True, exist_ok=True)
        args.profile_output.write_text(
            json.dumps(profile_for_json(profile), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    load_result = None
    if args.database_url:
        with psycopg2.connect(args.database_url) as conn:
            load_result = load_profile(conn, profile)

    summary = {
        "source_name": profile["source_name"],
        "source_checksum": profile["source_checksum"],
        "sheet_count": len(profile["sheets"]),
        "meaningful_rows": sum(sheet["meaningful_rows"] for sheet in profile["sheets"]),
        "staged_rows": sum(len(sheet["rows"]) for sheet in profile["sheets"]),
        "load_result": load_result,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
