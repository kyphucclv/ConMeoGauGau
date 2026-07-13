"""Audit transfer and mid-run join evidence in the legacy workbook."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = " ".join(str(value).strip().split())
    if not text or text.startswith("="):
        return None
    return text


def clean_code(value: Any) -> str | None:
    text = clean_text(value)
    if text and text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def parse_int(value: Any) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean_text(value)
    if text and text.isdigit():
        return int(text)
    return None


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def sheet_records(wb, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = []
    seen = defaultdict(int)
    for idx, value in enumerate(rows[0], start=1):
        base = clean_text(value) or f"column_{idx}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    records = []
    for row_number, values in enumerate(rows[1:], start=2):
        if not any(value is not None for value in values):
            continue
        record = {headers[idx]: value for idx, value in enumerate(values[: len(headers)])}
        record["_source_row_number"] = row_number
        records.append(record)
    return records


def audit(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    enrollments = []
    attendance_by_identity: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)

    for row in sheet_records(wb, "sheet2"):
        emp_code = clean_code(row.get("Emp Code"))
        class_code = clean_code(row.get("Class Code"))
        course_name = clean_text(row.get("Course Name"))
        if not emp_code or not class_code or not course_name:
            continue
        enrollments.append(
            {
                "source_row_number": row["_source_row_number"],
                "emp_code": emp_code,
                "class_code": class_code,
                "course_name": course_name,
                "start_date": parse_datetime(row.get("start date")),
                "final_level": clean_text(row.get("Final Level")),
                "entrance_level": clean_text(row.get("Entrance Level")),
            }
        )

    for row in sheet_records(wb, "ATTENDANCE_LOG"):
        emp_code = clean_code(row.get("Emp Code"))
        class_code = clean_code(row.get("Class Code"))
        course_name = clean_text(row.get("Course Name"))
        session_order = parse_int(row.get("Session Order"))
        dt = parse_datetime(row.get("Date"))
        if not emp_code or not class_code or not course_name:
            continue
        attendance_by_identity[(emp_code, class_code, course_name)].append(
            {
                "source_row_number": row["_source_row_number"],
                "session_order": session_order,
                "date": dt,
                "status": clean_text(row.get("Status")),
            }
        )

    by_emp: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for enrollment in enrollments:
        rows = attendance_by_identity[
            (enrollment["emp_code"], enrollment["class_code"], enrollment["course_name"])
        ]
        dated = [row for row in rows if row["date"]]
        sessions = [row["session_order"] for row in rows if row["session_order"]]
        enrollment["attendance_rows"] = len(rows)
        enrollment["first_attendance_date"] = min((row["date"] for row in dated), default=None)
        enrollment["last_attendance_date"] = max((row["date"] for row in dated), default=None)
        enrollment["first_session_order"] = min(sessions) if sessions else None
        enrollment["last_session_order"] = max(sessions) if sessions else None
        by_emp[enrollment["emp_code"]].append(enrollment)

    midrun_candidates = []
    no_attendance_enrollments = []
    for enrollment in enrollments:
        if enrollment["attendance_rows"] == 0:
            no_attendance_enrollments.append(enrollment)
            continue
        if enrollment["first_session_order"] and enrollment["first_session_order"] > 1:
            midrun_candidates.append(enrollment)

    transfer_candidates = []
    multi_class_employees = []
    multi_course_employees = []
    for emp_code, rows in sorted(by_emp.items()):
        classes = sorted({row["class_code"] for row in rows})
        courses = sorted({row["course_name"] for row in rows})
        if len(classes) > 1:
            multi_class_employees.append({"emp_code": emp_code, "classes": classes, "rows": rows})
        if len(courses) > 1:
            multi_course_employees.append({"emp_code": emp_code, "courses": courses, "rows": rows})

        ordered = sorted(
            rows,
            key=lambda row: (
                row["first_attendance_date"] or row["start_date"] or datetime.max,
                row["class_code"],
                row["course_name"],
            ),
        )
        for previous, current in zip(ordered, ordered[1:]):
            if previous["class_code"] != current["class_code"]:
                transfer_candidates.append(
                    {
                        "emp_code": emp_code,
                        "from": {
                            "class_code": previous["class_code"],
                            "course_name": previous["course_name"],
                            "last_attendance_date": previous["last_attendance_date"],
                            "source_row_number": previous["source_row_number"],
                        },
                        "to": {
                            "class_code": current["class_code"],
                            "course_name": current["course_name"],
                            "first_attendance_date": current["first_attendance_date"],
                            "first_session_order": current["first_session_order"],
                            "source_row_number": current["source_row_number"],
                        },
                    }
                )

    def json_safe(row: dict[str, Any]) -> dict[str, Any]:
        safe = {}
        for key, value in row.items():
            if isinstance(value, datetime):
                safe[key] = value.isoformat()
            elif isinstance(value, list):
                safe[key] = [json_safe(item) if isinstance(item, dict) else item for item in value]
            elif isinstance(value, dict):
                safe[key] = json_safe(value)
            else:
                safe[key] = value
        return safe

    return {
        "workbook": str(workbook_path),
        "enrollment_rows": len(enrollments),
        "enrollments_with_attendance": sum(1 for row in enrollments if row["attendance_rows"] > 0),
        "enrollments_without_attendance": len(no_attendance_enrollments),
        "midrun_candidate_count": len(midrun_candidates),
        "midrun_candidates": [json_safe(row) for row in midrun_candidates[:100]],
        "multi_class_employee_count": len(multi_class_employees),
        "multi_class_employee_examples": [json_safe(row) for row in multi_class_employees[:50]],
        "multi_course_employee_count": len(multi_course_employees),
        "transfer_candidate_count": len(transfer_candidates),
        "transfer_candidates": [json_safe(row) for row in transfer_candidates[:100]],
        "no_attendance_enrollment_examples": [json_safe(row) for row in no_attendance_enrollments[:50]],
        "decision": {
            "recommended_rule": (
                "Set start_session_number from the first observed attendance session only when it is > 1. "
                "Do not mark prior sessions absent. Treat multi-class employee histories as transfer candidates "
                "requiring review unless source rows define the transfer explicitly."
            )
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    result = audit(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "enrollment_rows": result["enrollment_rows"],
                "enrollments_without_attendance": result["enrollments_without_attendance"],
                "midrun_candidate_count": result["midrun_candidate_count"],
                "multi_class_employee_count": result["multi_class_employee_count"],
                "transfer_candidate_count": result["transfer_candidate_count"],
                "recommended_rule": result["decision"]["recommended_rule"],
            },
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
