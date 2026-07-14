"""Phase 8 automated verification and UAT gate."""

from __future__ import annotations

import os
import subprocess
import sys
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from auth import authenticate, bootstrap_first_admin, hash_password
from db import create_pool
from migrate import apply_migrations
from reporting import monthly_review_data, monthly_review_summary, monthly_review_xlsx, proposed_monthly_actions
from scripts.phase4_integration_check import _database_url, recreate_database
from services import BusinessService, CommandError


DEFAULT_MAINTENANCE_URL = "postgresql://postgres@localhost:5432/postgres"
DEFAULT_TEST_DB = "english_class_p8_test"
PG_DUMP = Path(r"C:\Program Files\PostgreSQL\18\bin\pg_dump.exe")
PG_RESTORE = Path(r"C:\Program Files\PostgreSQL\18\bin\pg_restore.exe")


def one(conn, sql: str, params: tuple = ()):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params)
        return cur.fetchone()


def all_rows(conn, sql: str, params: tuple = ()):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params)
        return cur.fetchall()


def expect_db_error(fn) -> None:
    try:
        fn()
    except psycopg2.Error:
        return
    raise AssertionError("expected PostgreSQL constraint error")


def expect_command_error(code: str, fn) -> None:
    try:
        fn()
    except CommandError as exc:
        if exc.code != code:
            raise AssertionError(f"expected {code}, got {exc.code}") from exc
        return
    raise AssertionError(f"expected CommandError {code}")


def verify_named_auth(database_url: str) -> int:
    conn = psycopg2.connect(database_url)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO app_users(username,password_hash,full_name,role)
                       VALUES('local_admin',%s,'Legacy Local Admin','admin')""",
                    (hash_password("unusable-local-password"),),
                )
    finally:
        conn.close()

    pool = create_pool(database_url, application_name="phase8_named_auth")
    try:
        user_id = bootstrap_first_admin(pool, "phase8_admin", "Phase 8 Admin", "admin-pass")
        user = authenticate(pool, "phase8_admin", "admin-pass")
        assert user is not None and user.user_id == user_id and user.role == "admin"
        assert authenticate(pool, "phase8_admin", "wrong-password") is None
        expect_command_error(
            "invalid_state",
            lambda: bootstrap_first_admin(pool, "second_admin", "Second Admin", "second-pass"),
        )
    finally:
        pool.closeall()

    conn = psycopg2.connect(database_url)
    try:
        audit = one(conn, "SELECT count(*) AS total FROM audit_events WHERE action='app_user.bootstrap_admin' AND actor_user_id=%s", (user_id,))
        assert audit["total"] == 1
    finally:
        conn.close()
    return user_id


def seed(conn, admin_user_id: int) -> dict[str, int]:
    with conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO app_users(username, password_hash, full_name, role)
                VALUES ('phase8_editor', %s, 'Phase 8 Editor', 'editor'),
                       ('phase8_viewer', %s, 'Phase 8 Viewer', 'viewer')
                RETURNING user_id, username
                """,
                (hash_password("editor-pass"), hash_password("viewer-pass")),
            )
            ids = {username: user_id for user_id, username in cur.fetchall()}
            ids["phase8_admin"] = admin_user_id
            cur.execute("INSERT INTO business_units(business_unit_name) VALUES('Original BU'), ('Updated BU') RETURNING business_unit_id, business_unit_name")
            ids.update({name.lower().replace(" ", "_"): row_id for row_id, name in cur.fetchall()})
            cur.execute("INSERT INTO job_roles(job_role_name) VALUES('Original Role'), ('Updated Role') RETURNING job_role_id, job_role_name")
            ids.update({name.lower().replace(" ", "_"): row_id for row_id, name in cur.fetchall()})
            cur.execute(
                """
                INSERT INTO courses(course_code, course_name, expected_units, attendance_threshold_ratio)
                VALUES ('UAT-A', 'UAT Course A', 4, 0.750),
                       ('UAT-B', 'UAT Course B', 4, 0.750)
                RETURNING course_id, course_code
                """
            )
            ids.update({f"course_{code[-1].lower()}": course_id for course_id, code in cur.fetchall()})
            cur.execute(
                """
                INSERT INTO levels(level_name, numeric_value, sequence_order)
                VALUES ('UAT Entrance', 1.0, 1),
                       ('UAT Middle', 2.0, 2),
                       ('UAT Peak', 3.0, 3)
                RETURNING level_id, level_name
                """
            )
            ids.update({name.lower().replace(" ", "_"): level_id for level_id, name in cur.fetchall()})
    return ids


def verify_migrations(conn) -> list[str]:
    versions = [row["version"] for row in all_rows(conn, "SELECT version FROM schema_migrations ORDER BY version")]
    expected = [
        "001_canonical_schema_v3",
        "002_raw_staging_and_profile",
        "003_etl_source_row_outcomes",
        "004_canonical_etl_batches",
        "005_phase4_completion",
        "006_reporting_views",
        "007_session_occurrences_and_pic_labels",
        "008_phase11_learner_transactions",
        "009_phase11_monthly_review",
        "010_phase11_course_creation_history_fix",
        "011_phase11_operational_issues",
        "012_phase11_unknown_org_placeholders",
        "013_phase11_legacy_attendance_exceptions",
        "014_phase11_unknown_placement_placeholder",
        "015_phase11_unknown_placement_numeric_fix",
        "016_phase11_runtime_invariants",
        "017_phase11_enrollment_membership_snapshot_remediation",
        "018_phase13_makeup_replacement_credit",
        "019_phase13_makeup_link_immutability",
    ]
    assert versions == expected
    return versions


def verify_schema_constraints(conn, svc: BusinessService, ids: dict[str, int]) -> None:
    employee = svc.create_or_update_employee("UAT-CONSTRAINT", "Constraint Learner").entity_id
    cohort = svc.create_cohort("UAT-CONSTRAINT", "Constraint Cohort").entity_id
    membership = svc.add_membership(cohort, employee, date(2026, 8, 1)).entity_id
    expect_command_error("duplicate", lambda: svc.add_membership(cohort, employee, date(2026, 8, 2)))

    run = svc.create_course_run(cohort, ids["course_a"], start_date=date(2026, 8, 1)).entity_id
    meeting = svc.save_meeting(run, datetime(2026, 8, 1, 9, 0, tzinfo=timezone.utc), 120).entity_id
    svc.add_session_unit(run, meeting, 1, unit_number_in_meeting=1)
    svc.add_session_unit(run, meeting, 2, unit_number_in_meeting=2)
    expect_command_error("invalid_state", lambda: svc.add_session_unit(run, meeting, 3, unit_number_in_meeting=3))

    atomic_cohort = svc.create_cohort("UAT-ATOMIC", "Atomic schedule").entity_id
    atomic_run = svc.create_course_run(atomic_cohort, ids["course_a"], start_date=date(2026, 8, 2)).entity_id
    occupied_meeting = svc.save_meeting(atomic_run, datetime(2026, 8, 2, 9, 0, tzinfo=timezone.utc), 60).entity_id
    svc.add_session_unit(atomic_run, occupied_meeting, 2)
    target_meeting = svc.save_meeting(atomic_run, datetime(2026, 8, 2, 11, 0, tzinfo=timezone.utc), 120).entity_id
    svc.add_session_unit(atomic_run, target_meeting, 99, unit_number_in_meeting=2)
    expect_command_error(
        "invalid_state",
        lambda: svc.add_session_units(atomic_run, target_meeting, 1, unit_count=2),
    )
    atomic_rows = one(conn, "SELECT count(*) AS total FROM session_units WHERE course_run_id=%s AND sequence_in_run=1", (atomic_run,))
    assert atomic_rows["total"] == 0
    svc.close_membership(membership, date(2026, 8, 31))


def run_uat(conn, ids: dict[str, int]) -> dict[str, object]:
    editor = BusinessService(conn, ids["phase8_editor"])
    admin = BusinessService(conn, ids["phase8_admin"])
    viewer = BusinessService(conn, ids["phase8_viewer"])

    expect_command_error("forbidden", lambda: viewer.create_cohort("UAT-VIEW", "Viewer Forbidden"))

    learner = editor.create_or_update_employee(
        "UAT-001",
        "UAT Learner",
        employment_status="active",
        business_unit_id=ids["original_bu"],
        job_role_id=ids["original_role"],
        valid_from=date(2026, 8, 1),
    ).entity_id
    pic = editor.create_or_update_employee("UAT-PIC", "UAT PIC", employment_status="active").entity_id
    midrun = editor.create_or_update_employee(
        "UAT-MID",
        "Mid Course Learner",
        employment_status="active",
        business_unit_id=ids["original_bu"],
        job_role_id=ids["original_role"],
        valid_from=date(2026, 8, 1),
    ).entity_id
    transfer = editor.create_or_update_employee(
        "UAT-TRANSFER",
        "Transfer Learner",
        employment_status="active",
        business_unit_id=ids["original_bu"],
        job_role_id=ids["original_role"],
        valid_from=date(2026, 8, 1),
    ).entity_id

    with conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO placements(employee_id, placement_kind, test_date, level_id) VALUES(%s, 'business', %s, %s)",
                (learner, date(2026, 8, 1), ids["uat_entrance"]),
            )

    cohort_a = editor.create_cohort("UAT-A", "UAT Cohort A", status="active").entity_id
    cohort_b = editor.create_cohort("UAT-B", "UAT Cohort B", status="active").entity_id
    editor.assign_pic(cohort_a, pic, date(2026, 8, 1))
    membership = editor.add_membership(cohort_a, learner, date(2026, 8, 1)).entity_id
    midrun_membership = editor.add_membership(cohort_a, midrun, date(2026, 8, 1)).entity_id
    transfer_membership = editor.add_membership(cohort_b, transfer, date(2026, 8, 1)).entity_id

    run_a = editor.create_course_run(cohort_a, ids["course_a"], start_date=date(2026, 8, 3)).entity_id
    run_b = editor.create_course_run(cohort_b, ids["course_a"], start_date=date(2026, 8, 5)).entity_id
    editor.change_course_run_status(run_a, "active")
    editor.change_course_run_status(run_b, "active")

    enrollment = editor.enroll(run_a, learner, membership, start_session_number=1).entity_id
    midrun_enrollment = editor.enroll(run_a, midrun, midrun_membership, start_session_number=2).entity_id
    transfer_enrollment = editor.enroll(run_b, transfer, transfer_membership, start_session_number=3).entity_id

    meeting = editor.save_meeting(run_a, datetime(2026, 8, 3, 9, 0, tzinfo=timezone.utc), 120, status="completed").entity_id
    cancelled = editor.save_meeting(run_a, datetime(2026, 8, 10, 9, 0, tzinfo=timezone.utc), 120).entity_id
    editor.cancel_meeting(cancelled, "UAT cancellation")
    cancelled_row = one(conn, "SELECT starts_at,duration_minutes,status,cancellation_reason FROM meetings WHERE meeting_id=%s", (cancelled,))
    assert cancelled_row["starts_at"] == datetime(2026, 8, 10, 9, 0, tzinfo=timezone.utc)
    assert cancelled_row["duration_minutes"] == 120
    assert cancelled_row["status"] == "cancelled"
    cancellation_audit = one(conn, "SELECT details FROM audit_events WHERE action='meeting.cancel' AND entity_key=%s ORDER BY created_at DESC LIMIT 1", (str(cancelled),))
    assert cancellation_audit["details"]["reason"] == "UAT cancellation"
    assert cancellation_audit["details"]["before"]["duration_minutes"] == 120
    assert cancellation_audit["details"]["after"]["status"] == "cancelled"
    final_test = editor.save_meeting(run_a, datetime(2026, 8, 17, 9, 0, tzinfo=timezone.utc), 180, status="completed").entity_id
    makeup_meeting = editor.save_meeting(run_a, datetime(2026, 8, 24, 9, 0, tzinfo=timezone.utc), 60, status="completed").entity_id

    unit_1 = editor.add_session_unit(run_a, meeting, 1, unit_number_in_meeting=1).entity_id
    unit_2 = editor.add_session_unit(run_a, meeting, 2, unit_number_in_meeting=2).entity_id
    cancelled_unit = editor.add_session_unit(run_a, cancelled, 3, unit_number_in_meeting=1).entity_id
    final_unit = editor.add_session_unit(run_a, final_test, 4, unit_number_in_meeting=1, unit_type="final_test").entity_id
    makeup_unit = editor.add_session_unit(run_a, makeup_meeting, 5, unit_number_in_meeting=1, unit_type="makeup").entity_id

    attendance = editor.bulk_record_attendance(
        [
            {"run_enrollment_id": enrollment, "session_unit_id": unit_1, "effective_status": "Present"},
            {"run_enrollment_id": enrollment, "session_unit_id": unit_2, "effective_status": "Absent"},
            {"run_enrollment_id": enrollment, "session_unit_id": cancelled_unit, "effective_status": "Absent"},
            {"run_enrollment_id": enrollment, "session_unit_id": final_unit, "effective_status": "Absent"},
            {"run_enrollment_id": midrun_enrollment, "session_unit_id": unit_2, "effective_status": "Present"},
        ]
    )
    makeup = editor.correct_attendance_makeup(attendance.values["attendance_ids"][1], makeup_unit, "UAT make-up")
    assert makeup.values["denominator_units_added"] == 0

    def rewrite_linked_makeup():
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE attendance SET session_unit_id=session_unit_id WHERE attendance_id=%s",
                    (makeup.entity_id,),
                )

    expect_db_error(rewrite_linked_makeup)
    expect_command_error(
        "duplicate_makeup",
        lambda: editor.correct_attendance_makeup(attendance.values["attendance_ids"][1], makeup_unit, "duplicate"),
    )

    attendance_row = one(conn, "SELECT * FROM v_run_enrollment_attendance WHERE run_enrollment_id=%s", (enrollment,))
    assert attendance_row["applicable_units"] == 3
    assert attendance_row["present_units"] == 2
    assert attendance_row["makeup_present_units"] == 1
    assert str(attendance_row["attendance_ratio"]) == "0.6667"
    assert attendance_row["calculated_exam_eligible"] is False
    admin.override_exam_eligibility(enrollment, True, "UAT reasoned override")
    expect_command_error(
        "invalid_input",
        lambda: editor.record_evaluation(enrollment, final_level_id=ids["uat_peak"], exam_eligible=True, passed=True),
    )

    midrun_row = one(conn, "SELECT applicable_units FROM v_run_enrollment_attendance WHERE run_enrollment_id=%s", (midrun_enrollment,))
    assert midrun_row["applicable_units"] == 2
    transfer_row = one(conn, "SELECT start_session_number FROM run_enrollments WHERE run_enrollment_id=%s", (transfer_enrollment,))
    assert transfer_row["start_session_number"] == 3

    editor.record_evaluation(
        enrollment,
        final_level_id=ids["uat_peak"],
        passed=True,
        next_course_id=ids["course_b"],
        teacher_notes="first evaluation",
        correction_reason="final result after approved eligibility override",
    )
    expect_command_error(
        "invalid_input",
        lambda: editor.record_evaluation(
            enrollment,
            final_level_id=ids["uat_middle"],
            passed=True,
            next_course_id=ids["course_b"],
            teacher_notes="correction without reason",
        ),
    )
    editor.record_evaluation(
        enrollment,
        final_level_id=ids["uat_middle"],
        passed=True,
        next_course_id=ids["course_b"],
        teacher_notes="correction creates regression",
        correction_reason="teacher corrected the final level",
    )
    no_continuation = editor.record_evaluation(
        transfer_enrollment,
        final_level_id=ids["uat_middle"],
        passed=True,
        next_course_id=None,
        teacher_notes="completed with no continuation",
    )
    assert no_continuation.entity_id
    editor.record_evaluation(midrun_enrollment, final_level_id=ids["uat_entrance"], passed=True, teacher_notes="first improvement fixture")
    editor.record_evaluation(midrun_enrollment, final_level_id=ids["uat_peak"], passed=True, teacher_notes="improved fixture", correction_reason="teacher confirmed improved level")

    suggestion = editor.suggest_completion(enrollment)
    assert suggestion.values["suggested"] is True
    admin.confirm_completion(enrollment, True)
    repeat_run = editor.create_course_run(cohort_a, ids["course_b"], start_date=date(2026, 8, 28)).entity_id
    editor.change_course_run_status(repeat_run, "active")
    editor.enroll(repeat_run, learner, membership, start_session_number=1)

    editor.create_or_update_employee(
        "UAT-001",
        "UAT Learner",
        employment_status="active",
        business_unit_id=ids["updated_bu"],
        job_role_id=ids["updated_role"],
        valid_from=date(2026, 9, 1),
    )
    snapshot = one(conn, "SELECT enrollment_business_unit, enrollment_job_role FROM v_historical_enrollment_snapshot WHERE run_enrollment_id=%s", (enrollment,))
    assert snapshot["enrollment_business_unit"] == "Original BU"
    assert snapshot["enrollment_job_role"] == "Original Role"

    progress = one(conn, "SELECT * FROM v_employee_progress_summary WHERE employee_id=%s", (learner,))
    assert progress["current_level_name"] == "UAT Middle"
    assert progress["highest_level_name"] == "UAT Peak"
    assert progress["regression_flag"] is True

    monthly = one(conn, "SELECT * FROM v_monthly_session_units WHERE course_run_id=%s", (run_a,))
    assert monthly["credited_session_units"] == 3
    assert monthly["final_test_units"] == 1

    before_issue_units = one(conn, "SELECT non_cancelled_units FROM v_cohort_course_run_dashboard WHERE course_run_id=%s", (run_a,))["non_cancelled_units"]
    with conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO data_quality_issues(issue_code, entity_type, entity_key, source_sheet, source_row_number, details)
                VALUES('legacy_anomaly_unresolved', 'meeting', 'UAT-A', 'ATTENDANCE_LOG', 88, '{}'::jsonb)
                RETURNING issue_id
                """
            )
            issue_id = cur.fetchone()[0]
    unresolved = one(conn, "SELECT count(*) AS total FROM v_unresolved_quality_issues WHERE issue_code='legacy_anomaly_unresolved'")
    assert unresolved["total"] == 1
    after_issue_units = one(conn, "SELECT non_cancelled_units FROM v_cohort_course_run_dashboard WHERE course_run_id=%s", (run_a,))["non_cancelled_units"]
    assert before_issue_units == after_issue_units
    editor.resolve_quality_issue(issue_id, "ignored", "accepted for UAT trace")

    eval_versions = one(conn, "SELECT count(*) AS total FROM evaluation_versions ev JOIN evaluations e ON e.evaluation_id=ev.evaluation_id WHERE e.run_enrollment_id=%s", (enrollment,))
    makeup_audit = one(conn, "SELECT count(*) AS total FROM audit_events WHERE action='attendance.makeup'")
    makeup_detail = one(conn, "SELECT details FROM audit_events WHERE action='attendance.makeup' ORDER BY audit_event_id DESC LIMIT 1")
    original_absence = one(conn, "SELECT effective_status FROM attendance WHERE attendance_id=%s", (attendance.values["attendance_ids"][1],))
    assert eval_versions["total"] == 3
    assert makeup_audit["total"] >= 1
    assert makeup_detail["details"]["denominator_units_added"] == 0
    assert makeup_detail["details"]["before"]["original_status"] == "Absent"
    assert makeup_detail["details"]["after"]["credited_status"] == "Present"
    assert original_absence["effective_status"] == "Absent"

    return {
        "attendance_ratio": str(attendance_row["attendance_ratio"]),
        "midrun_applicable_units": midrun_row["applicable_units"],
        "transfer_start_session_number": transfer_row["start_session_number"],
        "credited_session_units": monthly["credited_session_units"],
        "final_test_units": monthly["final_test_units"],
        "evaluation_versions": eval_versions["total"],
        "historical_bu_snapshot": snapshot["enrollment_business_unit"],
        "regression_flag": progress["regression_flag"],
        "quality_issue_traced": True,
    }


def verify_monthly_review(conn, ids: dict[str, int], database_url: str) -> dict[str, object]:
    review_month = date(2026, 8, 1)
    with conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE evaluation_versions SET created_at='2026-08-25 12:00:00+00'")
    pool = create_pool(database_url, application_name="phase8_monthly_review")
    try:
        data = monthly_review_data(pool, review_month)
        summary = monthly_review_summary(data)
        assert summary["repeated"] >= 1
        assert summary["improved_count"] >= 1
        assert summary["delivery_rate"] is not None
        assert data["course_participation"]
        assert data["class_participation"]
        learner_attendance = next(row for row in data["participation"] if row["emp_code"] == "UAT-001")
        assert learner_attendance["applicable_sessions"] == 3
        assert learner_attendance["present_sessions"] == 2
        assert str(learner_attendance["attendance_ratio"]) == "0.6667"
        editor = BusinessService(conn, ids["phase8_editor"])
        draft = proposed_monthly_actions(summary)
        saved = editor.save_monthly_action_summary(review_month, **draft)
        assert saved.values["version_number"] == 1
        refreshed = monthly_review_data(pool, review_month)
        assert refreshed["action_summary"]["version_number"] == 1
        export = monthly_review_xlsx(review_month, refreshed, draft)
        assert export[:2] == b"PK"
        workbook = load_workbook(BytesIO(export), read_only=True)
        assert {"Course participation", "Class participation"}.issubset(set(workbook.sheetnames))
        return {"repeated_participants": summary["repeated"], "improved_latest_tests": summary["improved_count"], "monthly_export": True}
    finally:
        pool.closeall()


def verify_operational_issue_inbox(conn, ids: dict[str, int]) -> dict[str, object]:
    editor = BusinessService(conn, ids["phase8_editor"])
    employee_id = editor.create_or_update_employee("UAT-ISSUE", "Issue Fixture").entity_id
    detected = one(conn, "SELECT count(*) AS total FROM v_operational_data_issues WHERE entity_type='employee' AND entity_key=%s", (str(employee_id),))
    assert detected["total"] >= 2
    unknown_backfill = BusinessService(conn, ids["phase8_admin"]).backfill_unknown_org_profiles()
    assert unknown_backfill.values["employee_count"] >= 1
    editor.create_or_update_employee(
        "UAT-ISSUE", "Issue Fixture", business_unit_id=ids["original_bu"], job_role_id=ids["original_role"], valid_from=date(2026, 8, 1),
    )
    with conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO placements(employee_id,placement_kind,test_date,level_id) VALUES(%s,'business',%s,%s)",
                        (employee_id, date(2026, 8, 1), ids["uat_entrance"]))
    corrected = one(conn, "SELECT count(*) AS total FROM v_operational_data_issues WHERE entity_type='employee' AND entity_key=%s", (str(employee_id),))
    assert corrected["total"] == 0
    missing_roster = one(conn, "SELECT entity_key FROM v_operational_data_issues WHERE issue_code='incomplete_attendance_roster' LIMIT 1")
    assert missing_roster
    attendance_facts_before = one(conn, "SELECT count(*) AS total FROM attendance WHERE session_unit_id=%s", (int(missing_roster["entity_key"]),))
    exception = BusinessService(conn, ids["phase8_admin"]).approve_legacy_attendance_exception(
        int(missing_roster["entity_key"]), "UAT: original paper roster is unavailable"
    )
    assert exception.entity_id == int(missing_roster["entity_key"])
    remaining_roster = one(conn, "SELECT count(*) AS total FROM v_operational_data_issues WHERE issue_code='incomplete_attendance_roster' AND entity_key=%s", (missing_roster["entity_key"],))
    assert remaining_roster["total"] == 0
    attendance_facts_after = one(conn, "SELECT count(*) AS total FROM attendance WHERE session_unit_id=%s", (int(missing_roster["entity_key"]),))
    assert attendance_facts_after["total"] == attendance_facts_before["total"]
    audit = one(conn, "SELECT count(*) AS total FROM audit_events WHERE action='attendance.legacy_exception.approve' AND entity_key=%s", (missing_roster["entity_key"],))
    assert audit["total"] == 1
    bulk_run = one(conn, "SELECT course_run_id FROM run_enrollments WHERE status='active' ORDER BY course_run_id LIMIT 1")["course_run_id"]
    bulk_session = editor.create_attendance_session(bulk_run, datetime(2026, 8, 31, 9, 0, tzinfo=timezone.utc), 60, 6)
    editor.save_meeting(bulk_run, datetime(2026, 8, 31, 9, 0, tzinfo=timezone.utc), 60, meeting_id=bulk_session.values["meeting_id"], status="completed")
    bulk_issue = one(conn, "SELECT count(*) AS total FROM v_operational_data_issues WHERE issue_code='incomplete_attendance_roster' AND entity_key=%s", (str(bulk_session.entity_id),))
    assert bulk_issue["total"] == 1
    bulk = BusinessService(conn, ids["phase8_admin"]).approve_all_legacy_attendance_exceptions("UAT: archived attendance source is unavailable")
    assert bulk.values["session_count"] >= 1
    bulk_remaining = one(conn, "SELECT count(*) AS total FROM v_operational_data_issues WHERE issue_code='incomplete_attendance_roster' AND entity_key=%s", (str(bulk_session.entity_id),))
    assert bulk_remaining["total"] == 0
    bulk_audit = one(conn, "SELECT details FROM audit_events WHERE action='attendance.legacy_exception.approve' AND entity_key=%s ORDER BY created_at DESC LIMIT 1", (str(bulk_session.entity_id),))
    assert bulk_audit["details"]["bulk_approval"] is True
    placement_employee = editor.create_or_update_employee(
        "UAT-PLACEMENT", "Placement Fixture", business_unit_id=ids["original_bu"], job_role_id=ids["original_role"], valid_from=date(2026, 8, 1)
    ).entity_id
    placement_issue = one(conn, "SELECT count(*) AS total FROM v_operational_data_issues WHERE issue_code='missing_business_placement' AND entity_key=%s", (str(placement_employee),))
    assert placement_issue["total"] == 1
    placement_backfill = BusinessService(conn, ids["phase8_admin"]).backfill_unknown_business_placements()
    assert placement_backfill.values["employee_count"] >= 1
    placement = one(conn, """SELECT l.level_name,p.source_reference FROM placements p JOIN levels l ON l.level_id=p.level_id
                            WHERE p.employee_id=%s AND p.placement_kind='business'""", (placement_employee,))
    assert placement["level_name"] == "Unknown Entrance Level"
    assert placement["source_reference"]["source"] == "phase11_unknown_placement_placeholder"
    return {"operational_issue_detected": detected["total"], "unknown_org_backfilled": unknown_backfill.values["employee_count"], "legacy_attendance_exception": True, "bulk_legacy_attendance_exception": True, "unknown_placement_backfilled": placement_backfill.values["employee_count"], "operational_issue_corrected": True}


def streamlit_smoke(database_url: str) -> dict[str, int]:
    os.environ["APP_DATABASE_URL"] = database_url
    from streamlit.testing.v1 import AppTest

    app = AppTest.from_file(str(ROOT / "streamlit_app.py"), default_timeout=10)
    app.run(timeout=10)
    assert not app.exception
    assert any("English class HR workspace" in item.value for item in app.title)
    assert not app.tabs
    sign_in_buttons = sum(1 for button in app.button if button.label == "Sign in")
    assert sign_in_buttons == 1
    next(item for item in app.text_input if item.label == "Username").input("phase8_admin")
    next(item for item in app.text_input if item.label == "Password").input("admin-pass")
    next(button for button in app.button if button.label == "Sign in").click()
    app.run(timeout=10)
    assert not app.exception
    assert [tab.label for tab in app.tabs] == [
        ":material/home_work: HR workspace",
        ":material/table_chart: Reports",
        ":material/history: Audit",
    ]
    assert any(button.label == "Sign out" for button in app.button)
    assert len(app.segmented_control) >= 1

    return {
        "titles": len(app.title),
        "tabs": len(app.tabs),
        "sign_in_buttons": sign_in_buttons,
    }


def backup_restore_rehearsal(database_url: str, restored_db: str, maintenance_url: str) -> dict[str, int]:
    if not PG_DUMP.exists() or not PG_RESTORE.exists():
        raise RuntimeError("PostgreSQL backup tools are not installed at the expected path")
    backup_path = ROOT / "backups" / "phase8_uat_rehearsal.dump"
    backup_path.parent.mkdir(exist_ok=True)
    if backup_path.exists():
        backup_path.unlink()

    subprocess.run([str(PG_DUMP), "--format=custom", "--file", str(backup_path), database_url], check=True)
    recreate_database(maintenance_url, restored_db)
    restored_url = _database_url(restored_db, maintenance_url)
    subprocess.run([str(PG_RESTORE), "--dbname", restored_url, str(backup_path)], check=True)

    source_conn = psycopg2.connect(database_url)
    restored_conn = psycopg2.connect(restored_url)
    try:
        source_counts = one(source_conn, "SELECT count(*) AS employees FROM employees")
        restored_counts = one(restored_conn, "SELECT count(*) AS employees FROM employees")
        assert source_counts["employees"] == restored_counts["employees"]
        restored_schema = one(restored_conn, "SELECT count(*) AS versions FROM schema_migrations")
        assert restored_schema["versions"] == 19
        return {"restored_employees": restored_counts["employees"], "restored_migrations": restored_schema["versions"]}
    finally:
        source_conn.close()
        restored_conn.close()


def main() -> None:
    db_name = os.getenv("PHASE8_TEST_DB", DEFAULT_TEST_DB)
    restored_db = f"{db_name}_restore"
    maintenance_url = os.getenv("PHASE8_MAINTENANCE_URL", DEFAULT_MAINTENANCE_URL)
    database_url = os.getenv("PHASE8_DATABASE_URL", _database_url(db_name, maintenance_url))

    recreate_database(maintenance_url, db_name)
    apply_migrations(database_url)
    admin_user_id = verify_named_auth(database_url)
    conn = psycopg2.connect(database_url)
    try:
        ids = seed(conn, admin_user_id)
        versions = verify_migrations(conn)
        editor = BusinessService(conn, ids["phase8_editor"])
        verify_schema_constraints(conn, editor, ids)
        uat = run_uat(conn, ids)
        monthly_review = verify_monthly_review(conn, ids, database_url)
        operational_issues = verify_operational_issue_inbox(conn, ids)
    finally:
        conn.close()

    smoke = streamlit_smoke(database_url)
    restore = backup_restore_rehearsal(database_url, restored_db, maintenance_url)

    print("Phase 8 automated verification and UAT gate passed.")
    print(f"migrations: {len(versions)}")
    for key, value in uat.items():
        print(f"{key}: {value}")
    for key, value in monthly_review.items():
        print(f"{key}: {value}")
    for key, value in operational_issues.items():
        print(f"{key}: {value}")
    for key, value in smoke.items():
        print(f"{key}: {value}")
    for key, value in restore.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
