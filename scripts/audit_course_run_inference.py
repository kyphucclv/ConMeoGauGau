"""Audit whether legacy workbook evidence supports course-run splitting."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = " ".join(str(value).strip().split())
    if not text or text.startswith("="):
        return None
    return text


def clean_code(value: Any) -> str | None:
    text = clean_text(value)
    if text and text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def parse_int(value: Any) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = clean_text(value)
    if text and text.isdigit():
        return int(text)
    return None


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def sheet_records(wb, sheet_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = []
    seen: Counter[str] = Counter()
    for idx, value in enumerate(rows[0], start=1):
        base = clean_text(value) or f"column_{idx}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")

    records = []
    for row_number, values in enumerate(rows[1:], start=2):
        if not any(value is not None for value in values):
            continue
        record = {headers[idx]: value for idx, value in enumerate(values[: len(headers)])}
        record["_source_row_number"] = row_number
        records.append(record)
    return records


def summarize_pair(attendance_rows: list[dict[str, Any]]) -> dict[str, Any]:
    dates = [row["date"] for row in attendance_rows if row["date"]]
    session_orders = [row["session_order"] for row in attendance_rows if row["session_order"]]
    date_by_session: dict[int, set[str]] = defaultdict(set)
    sessions_by_date: dict[str, set[int]] = defaultdict(set)
    for row in attendance_rows:
        session_order = row["session_order"]
        dt = row["date"]
        if session_order and dt:
            key = dt.isoformat()
            date_by_session[session_order].add(key)
            sessions_by_date[key].add(session_order)

    conflicting_sessions = {
        session: sorted(values)
        for session, values in date_by_session.items()
        if len(values) > 1
    }
    overfull_meetings = {
        dt: sorted(values)
        for dt, values in sessions_by_date.items()
        if len(values) > 2
    }
    session_date_counts = {
        str(session): len(values)
        for session, values in sorted(date_by_session.items())
        if len(values) > 1
    }
    reset_candidates = []
    previous_session = None
    previous_date = None
    unique_date_sessions = sorted(
        {
            (r["date"], r["session_order"])
            for r in attendance_rows
            if r["date"] and r["session_order"]
        }
    )
    for dt, session_order in unique_date_sessions:
        if previous_session is not None and dt > previous_date and session_order < previous_session:
            reset_candidates.append(
                {
                    "date": dt.isoformat(),
                    "session_order": session_order,
                    "previous_date": previous_date.isoformat(),
                    "previous_session_order": previous_session,
                }
            )
        if previous_date is None or dt > previous_date or session_order > previous_session:
            previous_session = session_order
            previous_date = dt

    return {
        "attendance_rows": len(attendance_rows),
        "distinct_learners": len({row["emp_code"] for row in attendance_rows if row["emp_code"]}),
        "date_min": min(dates).isoformat() if dates else None,
        "date_max": max(dates).isoformat() if dates else None,
        "session_min": min(session_orders) if session_orders else None,
        "session_max": max(session_orders) if session_orders else None,
        "distinct_session_orders": len(set(session_orders)),
        "session_orders_with_multiple_dates_count": len(session_date_counts),
        "session_orders_with_multiple_dates_examples": dict(list(session_date_counts.items())[:10]),
        "conflicting_session_count": len(conflicting_sessions),
        "conflicting_session_examples": dict(list(conflicting_sessions.items())[:10]),
        "overfull_meeting_count": len(overfull_meetings),
        "overfull_meeting_examples": dict(list(overfull_meetings.items())[:10]),
        "session_reset_candidate_count": len(reset_candidates),
        "session_reset_candidate_examples": reset_candidates[:10],
    }


def audit(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    class_dates = sheet_records(wb, "CLASS_DATES")
    sheet2 = sheet_records(wb, "sheet2")
    attendance = sheet_records(wb, "ATTENDANCE_LOG")

    pairs: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "class_dates_rows": [],
            "enrollment_rows": [],
            "attendance_rows": [],
        }
    )

    for row in class_dates:
        class_code = clean_code(row.get("Class Code"))
        course_name = clean_text(row.get("Course Name"))
        if class_code and course_name:
            pairs[(class_code, course_name)]["class_dates_rows"].append(row["_source_row_number"])

    for row in sheet2:
        class_code = clean_code(row.get("Class Code"))
        course_name = clean_text(row.get("Course Name"))
        emp_code = clean_code(row.get("Emp Code"))
        start_date = parse_datetime(row.get("start date"))
        if class_code and course_name:
            pairs[(class_code, course_name)]["enrollment_rows"].append(
                {
                    "source_row_number": row["_source_row_number"],
                    "emp_code": emp_code,
                    "start_date": start_date,
                }
            )

    for row in attendance:
        class_code = clean_code(row.get("Class Code"))
        course_name = clean_text(row.get("Course Name"))
        emp_code = clean_code(row.get("Emp Code"))
        session_order = parse_int(row.get("Session Order"))
        dt = parse_datetime(row.get("Date"))
        if class_code and course_name:
            pairs[(class_code, course_name)]["attendance_rows"].append(
                {
                    "source_row_number": row["_source_row_number"],
                    "emp_code": emp_code,
                    "session_order": session_order,
                    "date": dt,
                }
            )

    pair_summaries = []
    repeated_evidence = []
    ambiguous_evidence = []
    for (class_code, course_name), data in sorted(pairs.items()):
        summary = summarize_pair(data["attendance_rows"])
        enrollment_starts = sorted(
            {
                item["start_date"].date().isoformat()
                for item in data["enrollment_rows"]
                if item["start_date"]
            }
        )
        row = {
            "class_code": class_code,
            "course_name": course_name,
            "class_dates_count": len(data["class_dates_rows"]),
            "enrollment_rows": len(data["enrollment_rows"]),
            "enrollment_start_dates": enrollment_starts[:10],
            **summary,
        }
        pair_summaries.append(row)

        has_repeat_signal = (
            row["session_reset_candidate_count"] > 0
        )
        if has_repeat_signal:
            repeated_evidence.append(row)
        if row["conflicting_session_count"] or row["overfull_meeting_count"]:
            ambiguous_evidence.append(row)

    cohort_course_counts = Counter(class_code for class_code, _ in pairs)
    cohorts_with_many_courses = {
        class_code: count for class_code, count in sorted(cohort_course_counts.items()) if count > 1
    }

    return {
        "workbook": str(workbook_path),
        "pair_count": len(pair_summaries),
        "cohorts_with_many_courses": cohorts_with_many_courses,
        "repeated_run_candidate_count": len(repeated_evidence),
        "repeated_run_candidates": repeated_evidence,
        "ambiguous_pair_count": len(ambiguous_evidence),
        "ambiguous_pairs": ambiguous_evidence,
        "pair_summaries": pair_summaries,
        "decision": {
            "recommended_rule": (
                "Keep one course_run per class_code/course_name as Run 1 for the current ETL pass. "
                "The workbook does not expose a stable run identifier; repeated-session and "
                "conflicting-session evidence should remain issues until reviewed."
            )
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    result = audit(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "pair_count": result["pair_count"],
                "repeated_run_candidate_count": result["repeated_run_candidate_count"],
                "ambiguous_pair_count": result["ambiguous_pair_count"],
                "recommended_rule": result["decision"]["recommended_rule"],
            },
            indent=2,
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
